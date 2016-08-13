import unittest
from nose.tools import assert_equals
from nose.tools import assert_true
from selenium import webdriver
from selenium.webdriver.remote import webelement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.remote.webelement import WebElement
from openpyxl import load_workbook
from openpyxl import Workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from webstandards import webstandards
import unittest, time, re

class readandwriteio:

     def valueio(self):
        webstandard = webstandards()
        wb=load_workbook("C://Users//bharadwaj//test.xlsx",use_iterators=True)
        ws = wb.active
        self.wb1=Workbook()
        cellat=0
        for i in range(2,4):
            value=ws['A'+str(i)].value
            print value
            cellat=webstandard.io(value,i-1+int(cellat))

x=readandwriteio()
x.valueio()