#this class has all the UI components which can be accessed from different tests
'''
Author:braddy
Email:baryasom@asu.edu

8/1/2016
'''
import unittest
from nose.tools import assert_equals
from nose.tools import assert_true
from selenium import webdriver
from selenium.webdriver.remote import webelement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.remote.webelement import WebElement
from webstandarddirectory import webstandardsdirectory
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import unittest, time, re
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

class webstandards:
    #sheetvalue=writetoxl.io()
    def io(self,value,cellvalue):
        self.wb1 = load_workbook("textexample.xlsx")
        self.cellvalue=cellvalue
        self.sheet = self.wb1.active
        self.sheet['A1'+str(self.cellvalue)] = "URL for test"
        self.sheet['A'+str(self.cellvalue+1)] = "Tests"
        self.sheet['B'+str(self.cellvalue) ] = value
        self.sheet['B'+str(self.cellvalue+1)] = "Test-Status"
        self.sheet['C'+str(self.cellvalue+1)] = "Comments"
        self.cellvalue+=2
        self.cellvalue=self.global_header(value)
        self.wb1.save("textexample.xlsx")
        return self.cellvalue
    def access_url(self,url):
        self.url=url
        self.driver = webdriver.Remote(
   command_executor='http://127.0.0.1:4444/wd/hub',
   desired_capabilities=DesiredCapabilities.FIREFOX)
        global driver;
        driver = self.driver
        driver.get(url)
    #this would check if the header follows webstandards
    '''It would check for ASU logo size
        ASU links color
        ASU header background
        ASU main menu color before and after click
        ASU main menu font '''

    def global_header(self,url):
        cellvalue=self.cellvalue
        sheet=self.sheet
        self.access_url(url)
        directory = webstandardsdirectory()
        directory.asu_headerstandard()
        print "******* CHECKING ASU HEADER WEB STANDARDS********"
        self.linkcolor=directory.asuheader_linkcolor
        self.asubackground=directory.asuheader_backgroundcolor
        Image = driver.find_element_by_xpath("//img[@title='Arizona State University']")
        element = Image.size
        #asu logo
        if(element['width']!=203 or element['height']!=32):
            print "ASU LOGO not according to standard"
            sheet['A'+str(cellvalue)]="global header_logo"
            print "width and height are incorrect"+str(element['width']) +" "+ str(element['height'])
            sheet['B'+str(cellvalue)]="fail"
            sheet['C' + str(cellvalue)] =  "width and height are incorrect"+str(element['width']) +" "+ str(element['height'])
            cellvalue += 1
        else:
            sheet['A' + str(cellvalue)] = "global header_logo"
            print "global logo pass"
            sheet['B' + str(cellvalue)] = "pass"

            cellvalue += 1
        #asu_links
        asu_links=driver.find_elements_by_xpath("//a[@target='_top']")
        for i in asu_links:
            if len(str(i.text))>=1:
                color=i.value_of_css_property("color")
                if str(color)!=self.linkcolor:
                    sheet['A' + str(cellvalue)] = "global header_links_header"+i.text
                    print "links are iuncorrect"
                    sheet['B' + str(cellvalue)] = "fail"
                    sheet['C' + str(cellvalue)] = "NOT according to webstandards"+" "+str(color)+" element "+i.text
                    cellvalue += 1
                    print "NOT according to webstandards"+" "+str(color)+" element "+i.text
                else:
                    sheet['A' + str(cellvalue)] = "global header_links_header"+i.text
                    print "links are correct"
                    sheet['B' + str(cellvalue)] = "pass"
                    cellvalue += 1
                    print "according to webstandards" + " " + str(color) + " element " + i.text
        #asuheader-background
        asu_header=driver.find_element_by_xpath("//div[@id='asu_hdr']")
        asu_header_background=asu_header.value_of_css_property("background-color")
        if str(asu_header_background)!=self.asubackground:
            print "ASU HEADER background is not proper "
            sheet['A' + str(cellvalue)] = "global header_background "
            sheet['B' + str(cellvalue)] = "fail"
            sheet['C'+str(cellvalue)]="global header_background " + str(asu_header_background) +" is incorrect"
            cellvalue += 1
        #asuMainMenu
        try:
            asu_main_menu=driver.find_elements_by_xpath("//a[@href='#']")
            print "here"
            for i in range(1,len(asu_main_menu)-1):
                asu_mainmenucolor=asu_main_menu[i].value_of_css_property("color")
                asu_main_menu[i].click()
                WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.TAG_NAME, "footer")))
                hover=ActionChains(driver).move_to_element(asu_main_menu[i]).perform()
                asu_mainmenucolor_afterclick=asu_main_menu[i].value_of_css_property("color")
                asu_main_menufontfamily=asu_main_menu[i].value_of_css_property("font-family")
                asu_main_menufontsize = asu_main_menu[i].value_of_css_property("font-size")
                asu_main_menufontweight = asu_main_menu[i].value_of_css_property("font-weight")
                if str(asu_mainmenucolor)!=directory.maninmenucolor:
                    print str(asu_mainmenucolor)+" not according to webstandards "+asu_main_menu[i].text+ " "+directory.maninmenucolor
                if str(asu_mainmenucolor_afterclick )!= directory.mainmenucolor_click:
                    print str(asu_mainmenucolor_afterclick) +" not according to webstandards "+asu_main_menu[i].text+" "+directory.mainmenucolor_click
                if  str(asu_main_menufontfamily) != directory.fontmenu:
                    print str(asu_main_menufontfamily) + " not according to webstandards " + asu_main_menu[i].text+" "+directory.fontmenu
                if str(asu_main_menufontsize)!=directory.fontsize:
                    print str(asu_main_menufontsize) + " not according to webstandards " + asu_main_menu[i].text+" "+directory.fontsize
                if str(asu_main_menufontweight)!=directory.fontweight:
                    print str(asu_main_menufontweight) + " not according to webstandards " + asu_main_menu[i].text+" "+directory.fontweight
        except Exception:
            print "Main Menu was not accessed "
            sheet['A' + str(cellvalue)] = "global header_Mainmenu "
            sheet['B' + str(cellvalue)] = "NA"
            sheet['C' + str(cellvalue)] = "global header_Menu couldn't run"
            cellvalue += 1
        #seomobiletittle
        return cellvalue
        print "*********CHECKING ASU WEB STANDARDS COMPLETED******************"

    '''This would check font family on page is according to webstandards'''
    def asu_font_page(self,url):
        directory=webstandardsdirectory()
        directory.pagecontent_webstandards()
        self.font_family_standard=directory.font_family_standard
        self.primary_font=directory.primary_font_family_standard
        self.access_url(url)
        #WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "innovation-footer")))
        page_header_1=driver.find_elements_by_xpath("//h1")
        page_header_2 = driver.find_elements_by_xpath("//h2")
        page_header_3 = driver.find_elements_by_xpath("//h3")
        page_label=driver.find_elements_by_xpath("//label")
        page_paragraph=driver.find_elements_by_xpath("//p")
        page_option=driver.find_elements_by_xpath("//option")
        page_b=driver.find_elements_by_xpath("//b")

        print "******start Verifying font standards******"
        print "*******start verifying page labels"
        for i in page_label:
            font_family= i.value_of_css_property("font-family")
            font_family=str(font_family).replace('"','')
            if str(font_family).replace('"','')!=self.font_family_standard and str(font_family).replace('"','')!=self.primary_font :
                print str(i.text)+ "  is not according web standards"+" "+str(font_family)
        print "**********done verifying page standards*********"
        print "******start Verifying page header 1 font standards******"
        for j in page_header_1:
            font_family1 = j.value_of_css_property("font-family")
            if str(font_family1).replace('"','') != self.font_family_standard and str(font_family1).replace('"','')!=self.primary_font:
                print str(j.text) + "  is not according to  web standards"+str(font_family1)
        print "**********done verifying page standards*********"
        print "******start Verifying page header 2 font standards******"
        for k in page_header_2:
            font_family2 = k.value_of_css_property("font-family")
            if str(font_family2).replace('"','') != self.font_family_standard and str(font_family2).replace('"','')!=self.primary_font:
                print str(k.text) + "  is not according to  web standards"+str(font_family2)
        print "**********done verifying page standards*********"
        print "******start Verifying page header 3 font standards******"
        for l in page_header_3:
            font_family3 = l.value_of_css_property("font-family")
            if str(font_family3).replace('"','') != self.font_family_standard and str(font_family3).replace('"','')!=self.primary_font:
                print str(l.text) + " is not according to  web standards"+str(font_family3)
        print "**********done verifying page standards*********"
        print "******start Verifying page paragrapgh font standards******"
        for m in page_paragraph:
            font_family4 = m.value_of_css_property("font-family")
            if str(font_family4).replace('"','') != self.font_family_standard and str(font_family4).replace('"','')!=self.primary_font:
                print str(m.text) + " is not according to  web standards"+str(font_family4)
        print "**********done verifying page standards*********"
        print "******start Verifying page b font standards******"
        for n in page_b:
            font_family5 = n.value_of_css_property("font-family")
            if str(font_family5).replace('"','') != self.font_family_standard and str(font_family5).replace('"','')!=self.primary_font:
                print str(n.text) + " is not according to  web standards"+str(font_family5)
        print "**********done verifying page standards*********"
        print "done verifying font standards"+" elements verified are"

        #driver.close


    def asu_global_footer(self,url):
        self.access_url(url)
        directory = webstandardsdirectory()
        directory.asu_footerstandard()
        asuisno1="https://yourfuture.asu.edu/rankings"
        copyright="http://www.asu.edu/copyright/"
        accessibility="http://www.asu.edu/accessibility/"
        privacy="http://www.asu.edu/privacy/"
        jobs="http://www.asu.edu/asujobs"
        emergency="http://www.asu.edu/emergency/"
        contactasu="http://www.asu.edu/contactasu/"
        cssstyles=['font-size','font-weight','color','text-align']
        #asunoinnov=driver.find_element_by_xpath("//a[@href='"+asuisno1+"']")
        copyrig=driver.find_element_by_xpath("//a[@href='"+copyright+"']")
        accessible=driver.find_element_by_xpath("//a[@href='"+accessibility+"']")
        private=driver.find_element_by_xpath("//a[@href='"+privacy+"']")
        job=driver.find_element_by_xpath("//a[@href='"+jobs+"']")
        emerge=driver.find_element_by_xpath("//a[@href='"+emergency+"']")
        contact=driver.find_element_by_xpath("//a[@href='"+contactasu+"']")
        globalfooter=[copyrig,accessible,private,job,emerge,contact]
        for i in range(0,len(globalfooter)):
            cssvalue=globalfooter[i].value_of_css_property("font-size")
            if str(cssvalue)!=directory.fontsize_footer:
                print globalfooter[i].text +" "+cssvalue+" "+"is not according to webstandards"
            cssvalue = globalfooter[i].value_of_css_property("font-weight")
            if str(cssvalue) != directory.fontweight_footer:
                print globalfooter[i].text + " " + cssvalue + " " + "is not according to webstandards"
            cssvalue = globalfooter[i].value_of_css_property("color")
            if str(cssvalue) != directory.color_footer:
                print globalfooter[i].text + " " + cssvalue + " " + "is not according to webstandards"
            cssvalue = globalfooter[i].value_of_css_property("text-align")
            if str(cssvalue) != directory.textalign_footer:
                print globalfooter[i].text + " " + cssvalue + " " + "is not according to webstandards"

